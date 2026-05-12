import json
import os
from django.utils import timezone
from django.http import HttpResponse, FileResponse
from django.conf import settings
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from auth_app.models import User
from activity_log.utils import log_activity
from .models import Member, MembershipApplication
from .serializers import (
    MemberSerializer, MemberListSerializer,
    MembershipApplicationSerializer, MembershipApplicationListSerializer,
)


# ══════════════════════════════════════════════════════════════════
# HELPER — Auto-generate JSON + Excel after any member change
# ══════════════════════════════════════════════════════════════════

def sync_member_exports():
    """
    Called after every member add/edit/delete.
    Generates:
      - media/exports/members.json
      - media/exports/members.xlsx
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Get all members
        members = Member.objects.all().order_by('member_id')

        # Build data list
        data = []
        for m in members:
            data.append({
                'member_id':       m.member_id,
                'firstname':       m.firstname,
                'lastname':        m.lastname,
                'middlename':      m.middlename,
                'birthdate':       str(m.birthdate) if m.birthdate else '',
                'gender':          m.gender,
                'civil_status':    m.civil_status,
                'contact':         m.contact,
                'email':           m.email,
                'address':         m.address,
                'occupation':      m.occupation,
                'valid_id':        m.valid_id,
                'id_number':       m.id_number,
                'beneficiary':     m.beneficiary,
                'relationship':    m.relationship,
                'share_capital':   float(m.share_capital),
                'status':          m.status,
                'username':        m.user.username if m.user else '',
                'date_registered': m.date_registered.strftime('%Y-%m-%d %H:%M') if m.date_registered else '',
            })

        # Ensure exports directory exists
        export_dir = os.path.join(settings.MEDIA_ROOT, 'exports')
        os.makedirs(export_dir, exist_ok=True)

        # ── Save JSON ─────────────────────────────────────────────────────────
        json_path = os.path.join(export_dir, 'members.json')
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump({
                'generated_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
                'total':        len(data),
                'members':      data,
            }, f, indent=2, ensure_ascii=False)

        # ── Save Excel ────────────────────────────────────────────────────────
        xlsx_path = os.path.join(export_dir, 'members.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Members'

        # Header styling
        header_fill   = PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')
        header_font   = Font(bold=True, color='FFFFFF', size=11)
        header_align  = Alignment(horizontal='center', vertical='center')
        thin_border   = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        even_fill = PatternFill(start_color='F1F8E9', end_color='F1F8E9', fill_type='solid')

        # Title row
        ws.merge_cells('A1:S1')
        title_cell = ws['A1']
        title_cell.value       = f'LEAF MPC — Member List (Generated: {timezone.now().strftime("%B %d, %Y")})'
        title_cell.font        = Font(bold=True, size=13, color='1B5E20')
        title_cell.alignment   = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        # Column headers
        headers = [
            'Member ID','First Name','Last Name','Middle Name',
            'Birthdate','Gender','Civil Status','Contact','Email',
            'Address','Occupation','Valid ID','ID Number',
            'Beneficiary','Relationship','Share Capital',
            'Status','Username','Date Registered'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = thin_border
        ws.row_dimensions[2].height = 20

        # Data rows
        for row_num, m in enumerate(data, 3):
            row_data = [
                m['member_id'], m['firstname'], m['lastname'], m['middlename'],
                m['birthdate'], m['gender'], m['civil_status'], m['contact'], m['email'],
                m['address'], m['occupation'], m['valid_id'], m['id_number'],
                m['beneficiary'], m['relationship'], m['share_capital'],
                m['status'], m['username'], m['date_registered']
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border    = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                if row_num % 2 == 0:
                    cell.fill = even_fill

        # Column widths
        col_widths = [14,14,14,14,12,10,13,14,24,30,18,20,16,18,14,14,10,18,20]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Freeze header rows
        ws.freeze_panes = 'A3'

        wb.save(xlsx_path)

    except Exception as e:
        print(f'Export sync error: {e}')


# ══════════════════════════════════════════════════════════════════
# DOWNLOAD ENDPOINTS
# ══════════════════════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_members_json(request):
    if request.user.role not in ['admin', 'staff']:
        return Response({'error': 'Unauthorized.'}, status=403)

    json_path = os.path.join(settings.MEDIA_ROOT, 'exports', 'members.json')

    if not os.path.exists(json_path):
        sync_member_exports()

    with open(json_path, 'r', encoding='utf-8') as f:
        content = f.read()

    response = HttpResponse(content, content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename="LEAF_MPC_Members_{timezone.now().strftime("%Y%m%d")}.json"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_members_excel(request):
    if request.user.role not in ['admin', 'staff']:
        return Response({'error': 'Unauthorized.'}, status=403)

    xlsx_path = os.path.join(settings.MEDIA_ROOT, 'exports', 'members.xlsx')

    if not os.path.exists(xlsx_path):
        sync_member_exports()

    response = FileResponse(
        open(xlsx_path, 'rb'),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="LEAF_MPC_Members_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    return response


# ══════════════════════════════════════════════════════════════════
# MEMBERSHIP APPLICATIONS
# ══════════════════════════════════════════════════════════════════

@api_view(['GET', 'POST'])
def application_list_view(request):
    if request.method == 'GET':
        if not request.user.is_authenticated:
            return Response({'error': 'Authentication required.'}, status=401)
        if request.user.role not in ['admin', 'staff']:
            return Response({'error': 'Unauthorized.'}, status=403)
        apps = MembershipApplication.objects.all()
        if s := request.query_params.get('status'):
            apps = apps.filter(status=s)
        if q := request.query_params.get('search', '').strip():
            apps = apps.filter(lastname__icontains=q) | \
                   apps.filter(firstname__icontains=q) | \
                   apps.filter(app_id__icontains=q)
        return Response(MembershipApplicationListSerializer(apps, many=True).data)

    user = request.user if request.user.is_authenticated else None
    s = MembershipApplicationSerializer(data=request.data)
    if s.is_valid():
        app = s.save(user=user)
        log_activity('application', f'New online application received from {app.fullname} ({app.app_id})', user)
        return Response(MembershipApplicationSerializer(app).data, status=201)
    return Response(s.errors, status=400)


@api_view(['GET', 'PATCH'])
@permission_classes([IsAuthenticated])
def application_detail_view(request, pk):
    try:
        app = MembershipApplication.objects.get(pk=pk)
    except MembershipApplication.DoesNotExist:
        return Response({'error': 'Not found.'}, status=404)

    if request.method == 'GET':
        return Response(MembershipApplicationSerializer(app).data)

    if request.user.role not in ['admin', 'staff']:
        return Response({'error': 'Unauthorized.'}, status=403)

    new_status = request.data.get('status')
    if new_status in ['Approved', 'Rejected']:
        app.status      = new_status
        app.reviewed_at = timezone.now()
        app.reviewed_by = request.user.username
        if new_status == 'Rejected':
            app.reject_reason = request.data.get('reject_reason', '')
        app.save()
        log_activity('application', f'Application {app.app_id} ({app.fullname}) was {new_status.lower()} by {request.user.name}', request.user)
    return Response(MembershipApplicationSerializer(app).data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_application_view(request):
    try:
        app = MembershipApplication.objects.get(user=request.user)
        return Response(MembershipApplicationSerializer(app).data)
    except MembershipApplication.DoesNotExist:
        return Response({'error': 'No application found.'}, status=404)


# ══════════════════════════════════════════════════════════════════
# CONVERT APPLICATION → OFFICIAL MEMBER
# ══════════════════════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def convert_to_member_view(request, pk):
    if request.user.role not in ['admin', 'staff']:
        return Response({'error': 'Unauthorized.'}, status=403)

    try:
        app = MembershipApplication.objects.get(pk=pk)
    except MembershipApplication.DoesNotExist:
        return Response({'error': 'Application not found.'}, status=404)

    if app.status != 'Approved':
        return Response({'error': 'Application must be Approved first.'}, status=400)

    already_converted = Member.objects.filter(application=app).exists()
    if already_converted:
        existing = Member.objects.get(application=app)
        return Response({'error': 'Already converted.', 'member_id': existing.member_id}, status=400)

    import random, string
    user     = app.user
    plain_pw = ''

    if not user:
        fn = app.firstname.lower().strip()
        ln = app.lastname.lower().strip()
        base = f'{fn}.{ln}'; uname = base; i = 1
        while User.objects.filter(username=uname).exists():
            uname = f'{base}{i}'; i += 1
        plain_pw = 'leaf' + ''.join(random.choices(string.digits, k=4))
        user = User.objects.create_user(username=uname, password=plain_pw, name=f'{app.firstname} {app.lastname}', role='member')
        app.user = user; app.save()
    else:
        plain_pw = request.data.get('plain_password', '')
        user.role = 'member'; user.name = f'{app.firstname} {app.lastname}'.strip(); user.is_active = True; user.save()

    try:
        member = Member.objects.create(
            user=user, application=app,
            firstname=app.firstname, lastname=app.lastname, middlename=app.middlename,
            birthdate=app.birthdate, gender=app.gender, civil_status=app.civil_status,
            contact=app.contact, email=app.email, address=app.address,
            occupation=app.occupation, valid_id=app.valid_id, id_number=app.id_number,
            beneficiary=app.beneficiary, relationship=app.relationship,
            plain_password=plain_pw, status='Active',
        )
    except Exception as e:
        return Response({'error': f'Failed to create member: {str(e)}'}, status=500)

    sync_member_exports()  # ← Auto-update JSON + Excel
    log_activity('member', f'Member registered: {member.fullname} ({member.member_id}) — converted from {app.app_id}', request.user)

    return Response({
        'message': f'{member.fullname} is now an official member!',
        'member_id': member.member_id,
        'username': user.username,
        'member': MemberSerializer(member).data,
    }, status=201)


# ══════════════════════════════════════════════════════════════════
# OFFICIAL MEMBERS
# ══════════════════════════════════════════════════════════════════

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def member_list_view(request):
    if request.method == 'GET':
        members = Member.objects.all()
        if s := request.query_params.get('status'):
            members = members.filter(status=s)
        if q := request.query_params.get('search', '').strip():
            members = members.filter(lastname__icontains=q) | \
                      members.filter(firstname__icontains=q) | \
                      members.filter(member_id__icontains=q)
        return Response(MemberListSerializer(members, many=True).data)

    if request.user.role not in ['admin', 'staff']:
        return Response({'error': 'Unauthorized.'}, status=403)

    import random, string
    data          = request.data.copy()
    plain_password = data.get('plain_password', '') or 'leaf' + ''.join(random.choices(string.digits, k=4))

    fn = data.get('firstname', '').lower().strip()
    ln = data.get('lastname',  '').lower().strip()
    base = f'{fn}.{ln}'; uname = base; i = 1
    while User.objects.filter(username=uname).exists():
        uname = f'{base}{i}'; i += 1

    user = User.objects.create_user(
        username=uname, password=plain_password,
        name=f"{data.get('firstname','')} {data.get('lastname','')}".strip(),
        role='member', is_active=True,
    )

    try:
        member = Member.objects.create(
            user=user,
            firstname=data.get('firstname',''), lastname=data.get('lastname',''),
            middlename=data.get('middlename',''), birthdate=data.get('birthdate'),
            gender=data.get('gender','Male'), civil_status=data.get('civil_status','Single'),
            contact=data.get('contact',''), email=data.get('email',''),
            address=data.get('address',''), occupation=data.get('occupation',''),
            valid_id=data.get('valid_id',''), id_number=data.get('id_number',''),
            beneficiary=data.get('beneficiary',''), relationship=data.get('relationship',''),
            plain_password=plain_password, status='Active',
        )
    except Exception as e:
        user.delete()
        return Response({'error': f'Failed to register member: {str(e)}'}, status=500)

    sync_member_exports()  # ← Auto-update JSON + Excel
    log_activity('member', f'New member registered (F2F): {member.fullname} ({member.member_id}) by {request.user.name}', request.user)

    return Response({
        'message':        f'{member.fullname} is now an official member!',
        'member_id':      member.member_id,
        'username':       user.username,
        'plain_password': plain_password,
        'member':         MemberSerializer(member).data,
    }, status=201)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def member_detail_view(request, pk):
    try:
        member = Member.objects.get(pk=pk)
    except Member.DoesNotExist:
        return Response({'error': 'Not found.'}, status=404)

    if request.method == 'GET':
        return Response(MemberSerializer(member).data)

    if request.method == 'PUT':
        if request.user.role not in ['admin', 'staff']:
            return Response({'error': 'Unauthorized.'}, status=403)
        data   = request.data.copy()
        new_pw = data.get('plain_password', '')
        if new_pw and member.user:
            member.user.set_password(new_pw)
            member.user.save()
        s = MemberSerializer(member, data=data, partial=True)
        if s.is_valid():
            s.save()
            sync_member_exports()  # ← Auto-update JSON + Excel
            log_activity('member', f'Member updated: {member.fullname} ({member.member_id})', request.user)
            return Response(s.data)
        return Response(s.errors, status=400)

    if request.method == 'DELETE':
        if request.user.role != 'admin':
            return Response({'error': 'Unauthorized.'}, status=403)
        name = member.fullname
        mid  = member.member_id
        member.delete()
        sync_member_exports()  # ← Auto-update JSON + Excel (deleted member removed)
        log_activity('member', f'Member deleted: {name} ({mid})', request.user)
        return Response({'message': 'Member deleted.'})


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def member_status_view(request, pk):
    if request.user.role not in ['admin', 'staff']:
        return Response({'error': 'Unauthorized.'}, status=403)
    try:
        member = Member.objects.get(pk=pk)
    except Member.DoesNotExist:
        return Response({'error': 'Not found.'}, status=404)
    if st := request.data.get('status'):
        old = member.status; member.status = st; member.save()
        sync_member_exports()  # ← Auto-update JSON + Excel
        log_activity('member', f'Member status changed: {member.fullname} {old} → {st}', request.user)
    return Response(MemberSerializer(member).data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def member_stats_view(request):
    if request.user.role not in ['admin', 'staff']:
        return Response({'error': 'Unauthorized.'}, status=403)
    return Response({
        'total':                  Member.objects.count(),
        'active':                 Member.objects.filter(status='Active').count(),
        'inactive':               Member.objects.filter(status='Inactive').count(),
        'suspended':              Member.objects.filter(status='Suspended').count(),
        'pending_applications':   MembershipApplication.objects.filter(status='Pending').count(),
        'approved_applications':  MembershipApplication.objects.filter(status='Approved').count(),
        'total_applications':     MembershipApplication.objects.count(),
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_profile_view(request):
    try:
        member = Member.objects.get(user=request.user)
        return Response(MemberSerializer(member).data)
    except Member.DoesNotExist:
        return Response({'error': 'Not an official member yet.'}, status=404)