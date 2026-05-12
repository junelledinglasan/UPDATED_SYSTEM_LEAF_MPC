"""
LEAF MPC Report Exporters
Generates Excel and PDF reports from real database data.
"""
import io
from django.utils import timezone


# ══════════════════════════════════════════════════════════════════
# EXCEL EXPORTER
# ══════════════════════════════════════════════════════════════════

def generate_excel(report_type, date_from, date_to, data):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type[:31]

    # Styles
    GREEN       = '1B5E20'
    LIGHT_GREEN = 'E8F5E9'
    WHITE       = 'FFFFFF'
    GRAY        = 'F5F5F5'

    hdr_fill  = PatternFill(start_color=GREEN,       end_color=GREEN,       fill_type='solid')
    even_fill = PatternFill(start_color=LIGHT_GREEN,  end_color=LIGHT_GREEN, fill_type='solid')
    odd_fill  = PatternFill(start_color=GRAY,         end_color=GRAY,        fill_type='solid')
    hdr_font  = Font(bold=True, color=WHITE, size=11)
    title_font= Font(bold=True, color=GREEN, size=14)
    sub_font  = Font(color='555555', size=10)
    thin      = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    center    = Alignment(horizontal='center', vertical='center')
    left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

    cols = data.get('columns', [])
    rows = data.get('rows', [])
    ncols = max(len(cols), 1)

    # ── Title block ───────────────────────────────────────────────
    ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
    t = ws['A1']
    t.value     = '🌿  LEAF MPC — Cooperative Management System'
    t.font      = Font(bold=True, color=GREEN, size=15)
    t.alignment = center
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f'A2:{get_column_letter(ncols)}2')
    t2 = ws['A2']
    t2.value     = report_type
    t2.font      = title_font
    t2.alignment = center
    ws.row_dimensions[2].height = 24

    ws.merge_cells(f'A3:{get_column_letter(ncols)}3')
    t3 = ws['A3']
    t3.value     = f'Period: {date_from}  to  {date_to}   |   Generated: {timezone.now().strftime("%B %d, %Y %I:%M %p")}'
    t3.font      = sub_font
    t3.alignment = center
    ws.row_dimensions[3].height = 18

    # ── Summary section (if any) ──────────────────────────────────
    row_offset = 4
    summary = data.get('summary', [])
    if summary:
        ws.merge_cells(f'A{row_offset}:{get_column_letter(ncols)}{row_offset}')
        ws[f'A{row_offset}'].value = 'SUMMARY'
        ws[f'A{row_offset}'].font  = Font(bold=True, color=GREEN, size=11)
        ws[f'A{row_offset}'].alignment = center
        ws.row_dimensions[row_offset].height = 18
        row_offset += 1

        for item in summary:
            label, value = item
            ws[f'A{row_offset}'] = label
            ws[f'A{row_offset}'].font = Font(bold=True, size=10)
            ws.merge_cells(f'B{row_offset}:{get_column_letter(ncols)}{row_offset}')
            ws[f'B{row_offset}'] = value
            ws[f'B{row_offset}'].font = Font(size=10, color='1B5E20')
            row_offset += 1
        row_offset += 1

    # ── Column headers ────────────────────────────────────────────
    if cols:
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=row_offset, column=ci, value=col)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = center
            cell.border    = thin
        ws.row_dimensions[row_offset].height = 20
        row_offset += 1

    # ── Data rows ─────────────────────────────────────────────────
    for ri, row in enumerate(rows):
        fill = even_fill if ri % 2 == 0 else odd_fill
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=row_offset, column=ci, value=val)
            cell.fill      = fill
            cell.border    = thin
            cell.alignment = left_wrap
        ws.row_dimensions[row_offset].height = 16
        row_offset += 1

    # ── Auto column widths ────────────────────────────────────────
    widths = data.get('col_widths', [18] * ncols)
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = f'A{5 + len(summary) + (1 if summary else 0)}'

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════
# PDF EXPORTER
# ══════════════════════════════════════════════════════════════════

def generate_pdf(report_type, date_from, date_to, data):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buf = io.BytesIO()

    cols = data.get('columns', [])
    rows = data.get('rows', [])
    use_landscape = len(cols) > 6

    page_size = landscape(A4) if use_landscape else A4
    doc = SimpleDocTemplate(buf, pagesize=page_size,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm)

    GREEN     = colors.HexColor('#1B5E20')
    LT_GREEN  = colors.HexColor('#E8F5E9')
    MED_GREEN = colors.HexColor('#4CAF50')
    GRAY      = colors.HexColor('#F5F5F5')
    DARK      = colors.HexColor('#212121')
    WHITE     = colors.white

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title',
        fontSize=16, textColor=GREEN, fontName='Helvetica-Bold',
        alignment=TA_CENTER, spaceAfter=4)
    sub_style = ParagraphStyle('sub',
        fontSize=11, textColor=GREEN, fontName='Helvetica-Bold',
        alignment=TA_CENTER, spaceAfter=2)
    info_style = ParagraphStyle('info',
        fontSize=9, textColor=colors.gray,
        alignment=TA_CENTER, spaceAfter=8)
    label_style = ParagraphStyle('label',
        fontSize=10, textColor=DARK, fontName='Helvetica-Bold')
    val_style = ParagraphStyle('val',
        fontSize=10, textColor=GREEN)

    story = []

    # Header
    story.append(Paragraph('🌿  LEAF MPC — Cooperative Management System', title_style))
    story.append(Paragraph(report_type, sub_style))
    story.append(Paragraph(
        f'Period: {date_from}  to  {date_to}   |   Generated: {timezone.now().strftime("%B %d, %Y %I:%M %p")}',
        info_style))
    story.append(HRFlowable(width='100%', thickness=2, color=GREEN, spaceAfter=10))

    # Summary
    summary = data.get('summary', [])
    if summary:
        sum_data = [['Metric', 'Value']]
        for label, value in summary:
            sum_data.append([label, str(value)])

        sum_table = Table(sum_data, colWidths=[8*cm, 8*cm])
        sum_table.setStyle(TableStyle([
            ('BACKGROUND',  (0,0), (-1,0),  GREEN),
            ('TEXTCOLOR',   (0,0), (-1,0),  WHITE),
            ('FONTNAME',    (0,0), (-1,0),  'Helvetica-Bold'),
            ('FONTSIZE',    (0,0), (-1,0),  10),
            ('ALIGN',       (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME',    (0,1), (0,-1),  'Helvetica-Bold'),
            ('FONTSIZE',    (0,1), (-1,-1), 9),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [LT_GREEN, WHITE]),
            ('GRID',        (0,0), (-1,-1), 0.5, colors.HexColor('#C8E6C9')),
            ('TOPPADDING',  (0,0), (-1,-1), 5),
            ('BOTTOMPADDING',(0,0),(-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
        ]))
        story.append(sum_table)
        story.append(Spacer(1, 14))

    # Data table
    if cols and rows:
        page_w = landscape(A4)[0] - 3*cm if use_landscape else A4[0] - 3*cm
        col_widths = data.get('col_widths_pdf', None)
        if not col_widths:
            col_widths = [page_w / len(cols)] * len(cols)

        table_data = [cols] + [[str(v) if v is not None else '' for v in row] for row in rows]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,0),  GREEN),
            ('TEXTCOLOR',     (0,0), (-1,0),  WHITE),
            ('FONTNAME',      (0,0), (-1,0),  'Helvetica-Bold'),
            ('FONTSIZE',      (0,0), (-1,0),  9),
            ('ALIGN',         (0,0), (-1,0),  'CENTER'),
            ('FONTSIZE',      (0,1), (-1,-1), 8),
            ('FONTNAME',      (0,1), (-1,-1), 'Helvetica'),
            ('ROWBACKGROUNDS',(0,1), (-1,-1), [LT_GREEN, WHITE]),
            ('GRID',          (0,0), (-1,-1), 0.4, colors.HexColor('#C8E6C9')),
            ('TOPPADDING',    (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('LEFTPADDING',   (0,0), (-1,-1), 6),
            ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ]))
        story.append(table)

    # Footer
    story.append(Spacer(1, 16))
    story.append(HRFlowable(width='100%', thickness=1, color=MED_GREEN))
    story.append(Paragraph(
        f'LEAF MPC Cooperative Management System  •  Confidential  •  {timezone.now().strftime("%Y")}',
        ParagraphStyle('footer', fontSize=7, textColor=colors.gray, alignment=TA_CENTER, spaceBefore=4)
    ))

    doc.build(story)
    buf.seek(0)
    return buf